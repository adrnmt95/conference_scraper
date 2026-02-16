"""
Scrape economics/political science conferences from theeconomicmisfit.com
Uses OpenAI API to extract structured conference info from page text.
Filters out: macroeconomics, finance, pure economic/econometric theory
Exports to Excel sorted by closest application deadline.
Only processes new conferences not already in the Excel file.
"""

import requests
from bs4 import BeautifulSoup
import re
import json
import os
import warnings
from datetime import datetime, date
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
from dotenv import load_dotenv

warnings.filterwarnings("ignore", category=DeprecationWarning)

BASE_URL = "https://theeconomicmisfit.com/category/conferences/"
TODAY = date.today()
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")
XLSX_PATH = os.path.join(SCRIPT_DIR, "conferences.xlsx")

# Load environment variables from .env file if it exists (optional)
env_file = os.path.join(SCRIPT_DIR, ".env")
if os.path.exists(env_file):
    load_dotenv(env_file)

# Load config (only non-sensitive settings)
with open(CONFIG_PATH) as f:
    config = json.load(f)

# Get API key from environment variable (system-level)
openai_api_key = os.environ.get("OPENAI_API_KEY")
if not openai_api_key:
    raise ValueError("OPENAI_API_KEY not found in environment variables. Please set it with: export OPENAI_API_KEY='your_key'")

openai_client = OpenAI(api_key=openai_api_key)
OPENAI_MODEL = config.get("openai_model", "gpt-4o-mini")

# Persistent session with retries
session = requests.Session()
adapter = requests.adapters.HTTPAdapter(
    max_retries=requests.adapters.Retry(
        total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504]
    )
)
session.mount("https://", adapter)
session.mount("http://", adapter)
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
})

# Keywords that signal exclusion (macro, finance, pure theory)
EXCLUDE_KEYWORDS = [
    "finance phd workshop",
    "financial market",
    "household finance",
    "monetary economics",
    "monetary system",
    "asset pricing",
    "corporate finance",
    "econometric society",
    "econometrics and applied micro",
    "game and decision theory",
    "logic and the foundations of game",
    "ai in finance",
    "real estate",
    "macroeconomics and deep learning",
    "international monetary",
    "central bank research",
    "financial market research",
    "accounting group conference",
    "economic modeling and data science",
    "ecomod school",
    "ecomod2026",
    "nordic household finance",
]

# Keywords that signal inclusion (applied micro, labor, development, political economy, etc.)
INCLUDE_KEYWORDS = [
    "labor", "labour",
    "development",
    "political economy", "political science",
    "inequality",
    "conflict",
    "migration",
    "gender",
    "education",
    "health economics",
    "behavioral", "behavioural",
    "experimental economics",
    "urban economics",
    "environment",
    "energy",
    "agriculture",
    "public economics",
    "public policy",
    "policy evaluation", "policy impact",
    "infrastructure",
    "fragile",
    "humanitarian",
    "poverty",
    "africa",
    "latin america",
    "europe",
    "geopolitical",
    "defence economics",
    "defense economics",
    "organized crime",
    "philosophy, politics",
    "housing",
    "human capital",
    "demographic",
    "women",
    "opportunity",
    "jobs and development",
    "applied economics",
    "multidisciplinary",
]

# Titles to explicitly exclude
EXPLICIT_EXCLUDE_TITLES = [
    "9th dauphine finance",
    "nber summer institute",
    "pse-cepr policy forum",
    "sgf conference",
    "safe household finance",
    "nber-saif conference on ai and financial markets",
    "nordic household finance",
    "chapel hill-copenhagen conference on macroeconomics",
    "rcea international conference in economics, econometrics, and finance",
    "conference on new challenges in monetary economics",
    "11th annual meeting of the central bank research",
    "uncertainty, economic activity, and policy",
    "ai in finance conference",
    "baruch-nus real estate",
    "16th conference on logic and the foundations of game",
    "midwest international trade and theory",
    "third symposium on real estate",
    "annual meeting of the swiss society for financial market",
    "10th anniversary safe household finance",
    "eea and the econometric society european meeting",
    "2026 interactions conference",
    "the first chapel hill-copenhagen",
    "the 2026 rcea international conference",
    "ecomod school of modeling",
    "ecomod2026",
    "45th rsep international multidisciplinary",
    "gary chamberlain online seminar in econometrics",
    "submissions for the aea annual meeting",
    "symposium on forecasting",
    "ermass",
    "ersa2026 congress",
]


def load_existing_xlsx():
    """Load conferences from existing Excel file.
    Returns (known_titles set, active_rows list, past_rows list).
    - known_titles: normalized titles from both sheets (for skip check)
    - active_rows: conferences from "Conferences" sheet as dicts
    - past_rows: conferences from "Past Conferences" sheet as dicts
    """
    known_titles = set()
    active_rows = []
    past_rows = []

    if not os.path.exists(XLSX_PATH):
        return known_titles, active_rows, past_rows

    wb = load_workbook(XLSX_PATH, read_only=True)

    for sheet_name, target_list in [("Conferences", active_rows), ("Past Conferences", past_rows)]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for row in rows[1:]:
            row_dict = {}
            for col_idx, header in enumerate(headers):
                val = row[col_idx] if col_idx < len(row) else ""
                row_dict[header] = str(val).strip() if val else ""
            # Map Excel columns to internal keys
            conf = {
                "title": row_dict.get("Title", ""),
                "submission_deadline": row_dict.get("Submission Deadline", ""),
                "conference_dates": row_dict.get("Conference Dates", ""),
                "location": row_dict.get("Location", ""),
                "keynote_speakers": row_dict.get("Keynote Speakers", ""),
                "description": row_dict.get("Description", ""),
                "topics": row_dict.get("Topics", ""),
                "url": row_dict.get("URL", ""),
            }
            # Parse the deadline string back to a date object
            conf["deadline_date"] = parse_deadline_date(conf["submission_deadline"])
            if conf["title"]:
                known_titles.add(normalize_title(conf["title"]))
                target_list.append(conf)

    wb.close()
    return known_titles, active_rows, past_rows





def normalize_title(title):
    """Normalize a title for comparison (lowercase, strip non-alphanumeric)."""
    return re.sub(r"[^a-z0-9]", "", title.lower())


def get_all_conference_links():
    """Paginate through all listing pages and collect conference URLs."""
    all_links = []
    page = 1
    while True:
        url = BASE_URL if page == 1 else f"{BASE_URL}page/{page}/"
        print(f"Fetching listing page {page}: {url}")
        try:
            resp = session.get(url, timeout=60)
            if resp.status_code == 404:
                break
            resp.raise_for_status()
        except (requests.exceptions.HTTPError, requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            print(f"  Error on page {page}: {e}")
            if page > 1:
                time.sleep(5)
                try:
                    resp = session.get(url, timeout=90)
                    resp.raise_for_status()
                except Exception:
                    break
            else:
                break

        soup = BeautifulSoup(resp.text, "html.parser")

        links_found = set()
        for a_tag in soup.find_all("a", href=True):
            href = a_tag["href"]
            if re.match(
                r"https://theeconomicmisfit\.com/\d{4}/\d{2}/\d{2}/[\w-]+/?$", href
            ):
                if href not in links_found:
                    links_found.add(href)

        if not links_found:
            break

        all_links.extend(links_found)
        print(f"  Found {len(links_found)} conference links on page {page}")
        page += 1
        time.sleep(1)

    seen = set()
    unique = []
    for link in all_links:
        if link not in seen:
            seen.add(link)
            unique.append(link)

    print(f"\nTotal unique conference links: {len(unique)}")
    return unique


def fetch_page_text(url):
    """Fetch a conference page and return (title, full_text) or (None, None)."""
    try:
        resp = session.get(url, timeout=60)
        resp.raise_for_status()
    except Exception as e:
        print(f"  Error fetching {url}: {e}")
        return None, None

    soup = BeautifulSoup(resp.text, "html.parser")

    title_tag = soup.find("h1") or soup.find("h2", class_=re.compile("title"))
    title = title_tag.get_text(strip=True) if title_tag else ""

    content_div = (
        soup.find("div", class_=re.compile("entry-content"))
        or soup.find("div", class_=re.compile("post-content"))
        or soup.find("article")
    )

    if not content_div:
        return title, None

    full_text = content_div.get_text(separator="\n", strip=True)
    return title, full_text


def extract_with_openai(title, page_text):
    """Use OpenAI to extract structured conference info from page text."""
    prompt = f"""Extract the following fields from this conference announcement page.
Return a JSON object with exactly these keys. Use empty string "" if a field is not found.

- "submission_deadline": The submission/paper deadline as a human-readable string (e.g. "March 30, 2026")
- "deadline_date": The submission deadline as an ISO date YYYY-MM-DD (e.g. "2026-03-30"). If the year is missing, assume 2026.
- "conference_dates": When the conference takes place (e.g. "September 4-5, 2026")
- "location": Where the conference is held (city, country, or institution)
- "keynote_speakers": Names of keynote/invited/plenary speakers, comma-separated
- "description": A 1-2 sentence summary of what the conference is about
- "topics": Key topics or themes of interest, comma-separated

Conference title: {title}

Page text:
{page_text[:4000]}"""

    try:
        response = openai_client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "You extract structured data from conference announcements. Always respond with valid JSON only, no markdown fences."},
                {"role": "user", "content": prompt},
            ],
            temperature=0,
        )
        raw = response.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        return data
    except Exception as e:
        print(f"    OpenAI error: {e}")
        return {}


def parse_deadline_date(date_str):
    """Parse an ISO date string or common format into a date object."""
    if not date_str:
        return None
    # Try ISO format first
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        pass
    # Try common formats
    for fmt in ["%B %d, %Y", "%d %B %Y", "%B %d %Y"]:
        try:
            return datetime.strptime(date_str.replace(",", "").strip(), fmt).date()
        except ValueError:
            continue
    return None


def should_exclude(conf):
    """Decide whether a conference should be excluded based on topic."""
    title_lower = re.sub(r'["\u201c\u201d\u2018\u2019]', '', conf["title"].lower())
    full_lower = (conf.get("full_text") or conf.get("description", "") or "").lower()

    for excl in EXPLICIT_EXCLUDE_TITLES:
        if excl.lower() in title_lower:
            return True

    for kw in EXCLUDE_KEYWORDS:
        if kw in title_lower:
            return True

    combined = title_lower + " " + full_lower[:1500]
    has_include = any(kw in combined for kw in INCLUDE_KEYWORDS)

    strong_exclude_in_text = [
        "monetary policy", "central bank", "asset pricing",
        "financial econometrics", "corporate governance",
        "financial regulation", "portfolio", "stock market",
        "bond market", "banking regulation", "derivatives",
        "risk management in finance", "dsge", "new keynesian",
        "general equilibrium",
    ]

    strong_exclude_count = sum(1 for kw in strong_exclude_in_text if kw in combined)

    if strong_exclude_count >= 2 and not has_include:
        return True

    finance_title_words = ["finance", "monetary", "banking", "asset", "macroeconom"]
    title_finance_count = sum(1 for w in finance_title_words if w in title_lower)
    if title_finance_count >= 1 and not has_include:
        return True

    return False


def _format_deadline(conf):
    """Return a display string for the deadline."""
    dl = conf.get("deadline_date")
    if dl and not isinstance(dl, str):
        return dl.strftime("%B %d, %Y")
    if dl and isinstance(dl, str):
        return dl
    return conf.get("submission_deadline", "")


def _write_sheet(ws, conferences, header_fill, header_font, thin_border):
    """Write conference rows into a worksheet."""
    headers = [
        "Title", "Submission Deadline", "Conference Dates",
        "Location", "Keynote Speakers", "Description", "Topics", "URL",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    body_font = Font(name="Calibri", size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, conf in enumerate(conferences, 2):
        row_data = [
            conf.get("title", ""),
            _format_deadline(conf),
            conf.get("conference_dates", ""),
            conf.get("location", ""),
            conf.get("keynote_speakers", ""),
            conf.get("description", ""),
            conf.get("topics", ""),
            conf.get("url", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap_alignment
            cell.border = thin_border

    col_widths = [45, 22, 28, 35, 40, 60, 50, 55]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A2"


def write_to_excel(active_conferences, past_conferences, filename=None):
    """Write active and past conferences to an Excel file with two sheets."""
    filename = filename or XLSX_PATH
    wb = Workbook()

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Active conferences sheet
    ws_active = wb.active
    ws_active.title = "Conferences"
    _write_sheet(ws_active, active_conferences, header_fill, header_font, thin_border)

    # Past conferences sheet (keep only 10 most recent by deadline)
    ws_past = wb.create_sheet("Past Conferences")
    past_header_fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    _write_sheet(ws_past, past_conferences, past_header_fill, header_font, thin_border)

    wb.save(filename)
    print(f"\nExcel file saved: {filename}")


def main():
    print("=" * 60)
    print("Conference Scraper - theeconomicmisfit.com")
    print(f"Today's date: {TODAY}")
    print("=" * 60)

    # Load existing conferences from Excel (both active and past sheets)
    known_titles, existing_active, existing_past = load_existing_xlsx()
    print(f"\nAlready in Excel: {len(existing_active)} active, {len(existing_past)} past")

    # Check existing active conferences for passed deadlines -> move to past
    still_active = []
    newly_past = []
    for conf in existing_active:
        dl = conf.get("deadline_date")
        if dl and not isinstance(dl, str) and dl < TODAY:
            print(f"  Deadline passed: {conf['title'][:60]}")
            newly_past.append(conf)
        else:
            still_active.append(conf)

    # Merge newly past with existing past, sort by deadline (most recent first), keep 10
    all_past = existing_past + newly_past
    # Sort: those with a parsed deadline come first (most recent first), then those without
    all_past_with_dl = [c for c in all_past if c.get("deadline_date") and not isinstance(c["deadline_date"], str)]
    all_past_no_dl = [c for c in all_past if not c.get("deadline_date") or isinstance(c["deadline_date"], str)]
    all_past_with_dl.sort(key=lambda c: c["deadline_date"], reverse=True)
    past_conferences = (all_past_with_dl + all_past_no_dl)[:10]

    if newly_past:
        print(f"  Moved {len(newly_past)} conferences to Past Conferences")

    # Build set of known titles (active + past) for skipping
    active_title_set = {normalize_title(c["title"]) for c in still_active}

    # Step 1: Collect all conference links
    print("\n[1/4] Collecting conference links from all pages...")
    links = get_all_conference_links()

    # Step 2: Fetch page text and extract info (skip already-known titles)
    print(f"\n[2/4] Fetching details for {len(links)} conferences...")
    new_conferences = []
    new_count = 0
    skipped_count = 0

    for i, link in enumerate(links, 1):
        slug = link.split("/")[-2][:60]
        print(f"  [{i}/{len(links)}] {slug}")

        title, page_text = fetch_page_text(link)
        if not title:
            continue

        title_norm = normalize_title(title)

        # Skip if title is already in Excel (active or past)
        if title_norm in known_titles:
            print(f"    -> Already in Excel, skipping")
            skipped_count += 1
            continue

        if not page_text:
            continue

        # New conference: call OpenAI to extract fields
        print(f"    -> New conference, calling OpenAI...")
        extracted = extract_with_openai(title, page_text)

        deadline_date = parse_deadline_date(extracted.get("deadline_date", ""))

        conf = {
            "title": title,
            "url": link,
            "submission_deadline": extracted.get("submission_deadline", ""),
            "deadline_date": deadline_date,
            "conference_dates": extracted.get("conference_dates", ""),
            "location": extracted.get("location", ""),
            "keynote_speakers": extracted.get("keynote_speakers", ""),
            "description": extracted.get("description", ""),
            "topics": extracted.get("topics", ""),
            "full_text": page_text[:3000],
        }
        new_conferences.append(conf)
        new_count += 1

        time.sleep(0.5)

    print(f"\n  Scraped: {new_count} new, {skipped_count} skipped (already known)")

    # Step 3: Filter new conferences and merge with existing active
    print("\n[3/4] Filtering conferences...")
    filtered_new = []
    excluded_reasons = []

    for conf in new_conferences:
        dl = conf.get("deadline_date")
        if dl and not isinstance(dl, str) and dl < TODAY:
            excluded_reasons.append((conf["title"], "deadline passed"))
            # Still add to past if recent enough
            newly_past.append(conf)
            continue

        if should_exclude(conf):
            excluded_reasons.append((conf["title"], "topic excluded (macro/finance/theory)"))
            continue

        filtered_new.append(conf)

    # Combine existing active + filtered new
    all_active = still_active + filtered_new

    # Deduplicate by similar titles
    deduped = []
    for conf in all_active:
        norm = normalize_title(conf["title"])
        is_dup = False
        for j, existing_conf in enumerate(deduped):
            existing_norm = normalize_title(existing_conf["title"])
            shorter = min(len(norm), len(existing_norm))
            if shorter >= 15 and (norm.startswith(existing_norm[:shorter]) or existing_norm.startswith(norm[:shorter])):
                is_dup = True
                dl_new = conf.get("deadline_date")
                dl_old = existing_conf.get("deadline_date")
                if dl_new and not dl_old:
                    deduped[j] = conf
                elif len(conf.get("description", "")) > len(existing_conf.get("description", "")):
                    deduped[j] = conf
                break
        if not is_dup:
            deduped.append(conf)

    all_active = deduped

    # Sort: conferences with known deadlines first (by date), then those without
    with_deadline = [c for c in all_active if c.get("deadline_date") and not isinstance(c["deadline_date"], str)]
    without_deadline = [c for c in all_active if not c.get("deadline_date") or isinstance(c["deadline_date"], str)]

    with_deadline.sort(key=lambda c: c["deadline_date"])
    final_active = with_deadline + without_deadline

    # Rebuild past list (in case new scraped conferences also had passed deadlines)
    all_past = existing_past + newly_past
    all_past_with_dl = [c for c in all_past if c.get("deadline_date") and not isinstance(c["deadline_date"], str)]
    all_past_no_dl = [c for c in all_past if not c.get("deadline_date") or isinstance(c["deadline_date"], str)]
    all_past_with_dl.sort(key=lambda c: c["deadline_date"], reverse=True)
    # Deduplicate past by title
    seen_past = set()
    unique_past = []
    for c in all_past_with_dl + all_past_no_dl:
        tn = normalize_title(c["title"])
        if tn not in seen_past:
            seen_past.add(tn)
            unique_past.append(c)
    final_past = unique_past[:10]

    print(f"  Active: {len(final_active)} conferences")
    print(f"  Past (kept): {len(final_past)} conferences")
    print(f"  Excluded: {len(excluded_reasons)} conferences")
    if excluded_reasons:
        print("\n  Excluded conferences:")
        for title, reason in excluded_reasons:
            print(f"    - {title[:70]} [{reason}]")

    # Step 4: Save to Excel
    print(f"\n[4/4] Writing {len(final_active)} active + {len(final_past)} past conferences...")
    write_to_excel(final_active, final_past)

    print("\nDone!")
    print(f"  Active:  {len(final_active)} conferences")
    print(f"  Past:    {len(final_past)} conferences (10 most recent)")
    print(f"  XLSX:    {XLSX_PATH}")


if __name__ == "__main__":
    main()
