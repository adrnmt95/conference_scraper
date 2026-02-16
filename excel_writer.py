"""Excel writing and reading logic for conference data."""

import os
import re
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "conferences.xlsx")


def normalize_title(title):
    """Normalize a title for comparison (lowercase, strip non-alphanumeric)."""
    return re.sub(r"[^a-z0-9]", "", title.lower())


def parse_deadline_date(date_str):
    """Parse an ISO date string or common format into a date object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        pass
    for fmt in ["%B %d, %Y", "%d %B %Y", "%B %d %Y"]:
        try:
            return datetime.strptime(date_str.replace(",", "").strip(), fmt).date()
        except ValueError:
            continue
    return None


def load_existing_xlsx(xlsx_path=None):
    """Load conferences from existing Excel file.

    Returns (known_titles set, active_rows list, past_rows list, known_urls set).
    """
    xlsx_path = xlsx_path or XLSX_PATH
    known_titles = set()
    known_urls = set()
    active_rows = []
    past_rows = []

    if not os.path.exists(xlsx_path):
        return known_titles, active_rows, past_rows, known_urls

    wb = load_workbook(xlsx_path, read_only=True)

    for sheet_name, target_list in [("Conferences", active_rows), ("Past Conferences", past_rows)]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for row in rows[1:]:
            row_dict = {}
            for col_idx, header in enumerate(headers):
                val = row[col_idx] if col_idx < len(row) else ""
                row_dict[header] = str(val).strip() if val else ""
            conf = {
                "title": row_dict.get("Title", ""),
                "submission_deadline": row_dict.get("Submission Deadline", ""),
                "conference_dates": row_dict.get("Conference Dates", ""),
                "location": row_dict.get("Location", ""),
                "keynote_speakers": row_dict.get("Keynote Speakers", ""),
                "description": row_dict.get("Description", ""),
                "topics": row_dict.get("Topics", ""),
                "url": row_dict.get("URL", ""),
            }
            conf["deadline_date"] = parse_deadline_date(conf["submission_deadline"])
            if conf["title"]:
                known_titles.add(normalize_title(conf["title"]))
                target_list.append(conf)
            if conf["url"]:
                known_urls.add(conf["url"])

    wb.close()
    return known_titles, active_rows, past_rows, known_urls


def format_deadline(conf):
    """Return a display string for the deadline."""
    dl = conf.get("deadline_date")
    if dl and not isinstance(dl, str):
        return dl.strftime("%B %d, %Y")
    if dl and isinstance(dl, str):
        return dl
    return conf.get("submission_deadline", "")


def _write_sheet(ws, conferences, header_fill, header_font, thin_border):
    """Write conference rows into a worksheet."""
    headers = [
        "Title", "Submission Deadline", "Conference Dates",
        "Location", "Keynote Speakers", "Description", "Topics", "URL",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    body_font = Font(name="Calibri", size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, conf in enumerate(conferences, 2):
        row_data = [
            conf.get("title", ""),
            format_deadline(conf),
            conf.get("conference_dates", ""),
            conf.get("location", ""),
            conf.get("keynote_speakers", ""),
            conf.get("description", ""),
            conf.get("topics", ""),
            conf.get("url", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap_alignment
            cell.border = thin_border

    col_widths = [45, 22, 28, 35, 40, 60, 50, 55]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A2"


def write_to_excel(active_conferences, past_conferences, filename=None):
    """Write active and past conferences to an Excel file with two sheets."""
    filename = filename or XLSX_PATH
    wb = Workbook()

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_active = wb.active
    ws_active.title = "Conferences"
    _write_sheet(ws_active, active_conferences, header_fill, header_font, thin_border)

    ws_past = wb.create_sheet("Past Conferences")
    past_header_fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    _write_sheet(ws_past, past_conferences, past_header_fill, header_font, thin_border)

    wb.save(filename)
    print(f"\nExcel file saved: {filename}")
